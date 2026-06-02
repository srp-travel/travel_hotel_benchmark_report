#!/usr/bin/env python3
"""
generate_test_data.py — Fichier Excel de validation pour le Benchmark Tarifaire.
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet   # import direct — évite l'erreur Pylance

HOTEL = "Flamboyan Caribe Hotel & Spa"

ORX_HEADERS = [
    "Ville de départ", "Date de départ", "Nb jours", "Nb nuits",
    "Pension", "Catégorie", "Capacité de l'hébergement",
    "Type de prix", "Statut de disponibilité", "Statut global",
    "Prix NET TTC", "Comm. %", "Prix de vente TTC",
]

ORX_ROWS = [
    # Chambre double / Demi-Pension / 1* — 4 niveaux de compétitivité
    [None, "06/05/2026", 6, 5, "Demi-Pension", "Chambre double", None, "1*", "Disponible", "PUBLISHED", 212, 12, 240.00],
    [None, "07/05/2026", 6, 5, "Demi-Pension", "Chambre double", None, "1*", "Disponible", "PUBLISHED", 230, 12, 260.00],
    [None, "08/05/2026", 6, 5, "Demi-Pension", "Chambre double", None, "1*", "Disponible", "PUBLISHED", 240, 12, 272.00],
    [None, "09/05/2026", 6, 5, "Demi-Pension", "Chambre double", None, "1*", "Disponible", "PUBLISHED", 285, 12, 323.86],
    [None, "10/05/2026", 6, 5, "Demi-Pension", "Chambre double", None, "1*", "Disponible", "PUBLISHED", 212, 12, 240.00],
    [None, "11/05/2026", 6, 5, "Demi-Pension", "Chambre double", None, "1*", "Disponible", "PUBLISHED", 230, 12, 260.00],
    [None, "12/05/2026", 6, 5, "Demi-Pension", "Chambre double", None, "1*", "Disponible", "PUBLISHED", 240, 12, 272.00],
    # Logement seul — teste Meal Plan "NA" dans BKG
    [None, "06/05/2026", 6, 5, "Logement seul", "Chambre double", None, "1*", "Disponible", "PUBLISHED", 190, 12, 215.00],
    [None, "07/05/2026", 6, 5, "Logement seul", "Chambre double", None, "1*", "Disponible", "PUBLISHED", 190, 12, 215.00],
    # Chambre double vue mer
    [None, "06/05/2026", 6, 5, "Demi-Pension", "Chambre double vue mer", None, "1*", "Disponible", "PUBLISHED", 265, 12, 300.00],
    [None, "07/05/2026", 6, 5, "Demi-Pension", "Chambre double vue mer", None, "1*", "Disponible", "PUBLISHED", 265, 12, 300.00],
    # Suite — N/A intentionnel (pas de Room Type BKG)
    [None, "06/05/2026", 6, 5, "Demi-Pension", "Chambre Suite", None, "1*", "Disponible", "PUBLISHED", 420, 12, 475.00],
    # Date absente du BKG — trou de scraping
    [None, "15/05/2026", 6, 5, "Demi-Pension", "Chambre double", None, "1*", "Disponible", "PUBLISHED", 285, 12, 323.86],
]

BKG_HEADERS = [
    "Hotel Name", "Check-in", "Check-out",
    "Room Type", "Price", "Nights",
    "Meal Plan", "Cancellation Policy",
]

BKG_ROWS = [
    [HOTEL, "06/05/2026", "11/05/2026", "Chambre Double",                      622, 5, "Petit-déjeuner compris dans le tarif", "Free cancellation"],
    [HOTEL, "06/05/2026", "11/05/2026", "Chambre Lits Jumeaux avec Accès Spa", 643, 5, "Petit-déjeuner compris dans le tarif", "Non-refundable"],
    [HOTEL, "07/05/2026", "12/05/2026", "Chambre Double",                      622, 5, "Petit-déjeuner compris dans le tarif", "Free cancellation"],
    [HOTEL, "07/05/2026", "12/05/2026", "Chambre Lits Jumeaux avec Accès Spa", 643, 5, "Petit-déjeuner compris dans le tarif", "Non-refundable"],
    [HOTEL, "08/05/2026", "13/05/2026", "Chambre Double",                      622, 5, "Petit-déjeuner compris dans le tarif", "Free cancellation"],
    [HOTEL, "08/05/2026", "13/05/2026", "Chambre Lits Jumeaux avec Accès Spa", 643, 5, "Petit-déjeuner compris dans le tarif", "Non-refundable"],
    [HOTEL, "09/05/2026", "14/05/2026", "Chambre Double",                      622, 5, "Petit-déjeuner compris dans le tarif", "Free cancellation"],
    [HOTEL, "09/05/2026", "14/05/2026", "Chambre Lits Jumeaux avec Accès Spa", 643, 5, "Petit-déjeuner compris dans le tarif", "Non-refundable"],
    [HOTEL, "10/05/2026", "15/05/2026", "Chambre Double",                      622, 5, "Petit-déjeuner compris dans le tarif", "Free cancellation"],
    [HOTEL, "10/05/2026", "15/05/2026", "Chambre Lits Jumeaux avec Accès Spa", 643, 5, "Petit-déjeuner compris dans le tarif", "Non-refundable"],
    [HOTEL, "11/05/2026", "16/05/2026", "Chambre Double",                      622, 5, "Petit-déjeuner compris dans le tarif", "Free cancellation"],
    [HOTEL, "12/05/2026", "17/05/2026", "Chambre Double",                      622, 5, "Petit-déjeuner compris dans le tarif", "Free cancellation"],
    # Meal Plan = "NA" — logement seul
    [HOTEL, "06/05/2026", "11/05/2026", "Chambre Double",                      520, 5, "NA",                                   "Free cancellation"],
    [HOTEL, "07/05/2026", "12/05/2026", "Chambre Double",                      520, 5, "NA",                                   "Free cancellation"],
    # Chambre Vue Mer
    [HOTEL, "06/05/2026", "11/05/2026", "Chambre Double Vue Mer",              680, 5, "Petit-déjeuner compris dans le tarif", "Free cancellation"],
    [HOTEL, "07/05/2026", "12/05/2026", "Chambre Double Vue Mer",              680, 5, "Petit-déjeuner compris dans le tarif", "Free cancellation"],
    # Date hors ORX (extra BKG)
    [HOTEL, "20/06/2026", "25/06/2026", "Chambre Double",                      710, 5, "Petit-déjeuner compris dans le tarif", "Free cancellation"],
    # 15/05/2026 absent → trou de scraping
]


def _write_sheet(ws: Worksheet, headers: list[str], rows: list[list]) -> None:
    """Écrit les en-têtes et données dans une feuille avec mise en forme."""
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for ci, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(ci)].width = max(len(h) + 4, 16)
    ws.row_dimensions[1].height = 26

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell           = ws.cell(row=ri, column=ci, value=(val if val is not None else ""))
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def generate(output_path: str = "test_validation.xlsx") -> None:
    wb      = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    _write_sheet(wb.create_sheet("1. INPUT ORX EXPORT"), ORX_HEADERS, ORX_ROWS)
    _write_sheet(wb.create_sheet("2. OUTPUT BKG SCRAP"), BKG_HEADERS, BKG_ROWS)
    wb.save(output_path)

    print(f"✅ Fichier généré : {output_path}")
    print(f"   ORX : {len(ORX_ROWS)} lignes | BKG : {len(BKG_ROWS)} lignes")
    print()
    print("Résultats attendus :")
    print("  ✅ Très compétitif  — 240/pp × 2 = 480  vs 622  (-22.8%)")
    print("  ✅ Compétitif       — 260/pp × 2 = 520  vs 622  (-16.4%)")
    print("  ⚠️  Proche           — 272/pp × 2 = 544  vs 622  (-12.5%)")
    print("  ❌ Non compétitif   — 324/pp × 2 = 648  vs 622  (+4.1%)  ← données réelles")
    print("  ⬜ N/A (Suite)      — aucun Room Type BKG mappé")
    print("  ⬜ N/A (15/05)      — date absente du BKG")


if __name__ == "__main__":
    generate()