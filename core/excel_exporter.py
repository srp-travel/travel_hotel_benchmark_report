"""
excel_exporter.py — Génération du rapport Excel avec mise en forme conditionnelle.
Mis en cache via @st.cache_data pour éviter les recalculs inutiles.
"""

from __future__ import annotations

import io

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.constants import COL_WIDTHS, ROW_COLORS, SHEET_REPORT


@st.cache_data(show_spinner=False)
def build_excel(df_out: pd.DataFrame, orig_bytes: bytes, _df_full: pd.DataFrame) -> bytes:
    """
    Génère la feuille RAPPORT DETAILLE dans le classeur source.
    Paramètres :
      - df_out      : DataFrame sans la colonne _competitivite (colonnes d'affichage)
      - orig_bytes  : octets du fichier Excel original (pour conserver les autres feuilles)
      - _df_full    : DataFrame complet avec _competitivite (préfixé _ = ignoré par cache)
    """
    wb = openpyxl.load_workbook(io.BytesIO(orig_bytes))

    # Suppression de l'ancienne feuille si elle existe
    if SHEET_REPORT in wb.sheetnames:
        del wb[SHEET_REPORT]

    ws      = wb.create_sheet(SHEET_REPORT)
    headers = list(df_out.columns)

    # -- Styles en-tête
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin_bdr = Border(
        bottom=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin",  color="D0D0D0"),
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(h, 18)
    ws.row_dimensions[1].height = 32

    # -- Données
    for ri, (orig_idx, row_data) in enumerate(df_out.iterrows(), 2):
        comp     = str(_df_full.at[orig_idx, "_competitivite"])
        row_fill = PatternFill("solid", fgColor=ROW_COLORS.get(comp, ("F2F2F2", ""))[0])

        for ci, h in enumerate(headers, 1):
            val  = row_data[h]
            cell = ws.cell(row=ri, column=ci)
            cell.font   = Font(size=10)
            cell.border = thin_bdr
            cell.fill   = row_fill

            is_nan = val is None or (not isinstance(val, str) and pd.isna(val))
            if is_nan:
                cell.value = None
            elif h == "Ecart PCT" and isinstance(val, (int, float)):
                cell.value         = val
                cell.number_format = "0.0%"
                cell.alignment     = Alignment(horizontal="right")
            elif h in ("Ecart EUR", "Prix BKG (min)", "Prix ORX / chambre", "Prix de vente TTC") \
                    and isinstance(val, (int, float)):
                cell.value         = val
                cell.number_format = '#,##0.00 "EUR"'
                cell.alignment     = Alignment(horizontal="right")
            elif h in ("Date de départ", "Nb nuits"):
                cell.value     = val
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = val

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
