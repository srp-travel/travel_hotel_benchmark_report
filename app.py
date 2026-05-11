#!/usr/bin/env python3
"""
Benchmark Tarifaire ORX vs Booking.com
Usage : streamlit run benchmark_ui.py
Prérequis : pip install streamlit pandas openpyxl
"""

from __future__ import annotations

import io
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════

PENSIONS_NORM = [
    "PDJ - Petit-déjeuner",
    "DP - Demi-pension",
    "PC - Pension complète",
    "TI - Tout inclus",
    "LS - Logement seul",
]

ANNULATIONS_NORM = [
    "AG - Annulation gratuite",
    "NA - Non annulable",
    "NANR - Non annulable / Non remboursable",
    "NR - Non remboursable",
    "PR - Partiellement remboursable",
    "RF - Remboursable",
]

ROW_COLORS: dict[str, tuple[str, str]] = {
    "✅ Très compétitif": ("C6EFCE", "#C6EFCE"),
    "✅ Compétitif":       ("E2EFDA", "#E2EFDA"),
    "⚠️ Proche":           ("FFEB9C", "#FFEB9C"),
    "❌ Non compétitif":   ("FFC7CE", "#FFC7CE"),
    "N/A":                 ("F2F2F2", "#F2F2F2"),
}

COL_WIDTHS: dict[str, int] = {
    "Date de départ":                      16,
    "Nb nuits":                            10,
    "Catégorie ORX":                       22,
    "Pension ORX (norm.)":                 24,
    "Politique annulation ORX (réf.)":     34,
    "Type de prix":                        14,
    "Prix de vente TTC":                   18,
    "Prix ORX / chambre":                  18,
    "Room Type BKG":                       40,
    "Meal Plan BKG":                       36,
    "Meal Plan BKG (norm.)":               26,
    "Politique annulation BKG (norm.)":    38,
    "Prix BKG (min)":                      16,
    "Ecart EUR":                           13,
    "Ecart PCT":                           13,
}

PAGE_SIZES = [10, 25, 50, 100]

SEUIL_LABELS: list[tuple[str, str, str]] = [
    ("✅ Très compétitif", "success", "#C6EFCE"),
    ("✅ Compétitif",       "success", "#E2EFDA"),
    ("⚠️ Proche",           "warning", "#FFEB9C"),
    ("❌ Non compétitif",   "danger",  "#FFC7CE"),
]


# ══════════════════════════════════════════════════════════════
# PAGE & CSS
# ══════════════════════════════════════════════════════════════

st.set_page_config(page_title="Benchmark Tarifaire", page_icon="🏨", layout="wide")

st.markdown("""
<style>
    html, body, [class*="css"] { font-size: 15px; }
    .main .block-container { padding-top: 1.5rem; max-width: 1440px; }
    .step-title {
        display: flex; align-items: center; gap: 10px;
        font-size: 18px; font-weight: 700; color: #1F3864; margin: 0 0 8px 0;
    }
    .step-badge {
        display: inline-flex; align-items: center; justify-content: center;
        min-width: 28px; height: 28px; background: #1F3864; color: white;
        border-radius: 50%; font-size: 14px; font-weight: 700;
    }
    .bench-title  { font-size: 24px; font-weight: 700; color: #1F3864; margin-bottom: 3px; }
    .bench-subtitle { font-size: 15px; color: #555; margin-bottom: 16px; }
    .kpi-section-title {
        font-size: 13px; font-weight: 700; text-transform: uppercase;
        letter-spacing: .07em; color: #1F3864;
        border-bottom: 2px solid #1F3864; padding-bottom: 5px; margin: 14px 0 10px 0;
    }
    .kpi-row {
        display: flex; justify-content: space-between; align-items: center;
        padding: 6px 0; border-bottom: 1px solid #F0F0F0;
    }
    .kpi-label  { font-size: 14px; color: #444; }
    .kpi-value  { font-size: 15px; font-weight: 600; color: #1F3864; }
    .kpi-value.danger  { color: #C00000; }
    .kpi-value.warning { color: #9C6500; }
    .kpi-value.success { color: #375623; }
    .seuil-row {
        display: flex; justify-content: space-between; align-items: center;
        padding: 8px 10px; border-radius: 6px; margin-bottom: 6px;
    }
    .seuil-label { font-size: 14px; font-weight: 500; }
    .seuil-count { font-size: 15px; font-weight: 700; }
    .seuil-pct   { font-size: 13px; color: #555; margin-left: 6px; }
    .tw-bar-wrap {
        background: #E8E8E8; border-radius: 20px; height: 16px;
        overflow: hidden; margin: 8px 0 2px 0;
    }
    .tw-bar-fill { height: 100%; border-radius: 20px; transition: width .4s; }
    .legend-wrap  { display: flex; gap: 20px; flex-wrap: wrap; margin: 8px 0 14px 0; align-items: center; }
    .legend-item  { display: flex; align-items: center; gap: 7px; font-size: 14px; }
    .legend-dot   { width: 16px; height: 16px; border-radius: 4px; flex-shrink: 0; border: 1px solid #ccc; }
    .pagination-bar {
        display: flex; align-items: center; justify-content: center;
        gap: 14px; padding: 8px 0; font-size: 15px; color: #1F3864;
    }
    .stTabs [data-baseweb="tab-list"] { gap: 2px; }
    .stTabs [data-baseweb="tab"] { padding: 9px 20px; border-radius: 6px 6px 0 0; font-size: 14px; }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #1F3864 0%, #2E5FA3 100%) !important;
        color: #FFFFFF !important; border: none !important; border-radius: 8px !important;
        font-size: 15px !important; font-weight: 600 !important; letter-spacing: 0.04em !important;
        padding: 14px 28px !important; box-shadow: 0 3px 10px rgba(31,56,100,.35) !important;
        transition: all 0.2s ease !important; width: 100% !important;
    }
    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #162A4E 0%, #1F4A8A 100%) !important;
        box-shadow: 0 6px 18px rgba(31,56,100,.45) !important; transform: translateY(-2px) !important;
    }
    .stDownloadButton > button:active {
        transform: translateY(0px) !important; box-shadow: 0 2px 6px rgba(31,56,100,.3) !important;
    }
    .app-footer {
        text-align: center; padding: 24px 0 12px 0; font-size: 13px; color: #888;
        border-top: 1px solid #E8E8E8; margin-top: 40px;
    }
    .app-footer strong { color: #1F3864; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

st.markdown("# 🏨 Benchmark Tarifaire")
st.caption("Analyse de compétitivité tarifaire — ORX (Orchestra) vs Booking.com")
st.markdown("---")


# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════

def safe_unique(df: pd.DataFrame, col: str) -> list[str]:
    if col not in df.columns:
        return []
    return sorted(df[col].dropna().astype(str).str.strip().unique().tolist())


def normalize_price(val: object) -> float | None:
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        try:
            return float(val.replace(" ", "").replace("\u202f", "").replace(",", "."))
        except ValueError:
            return None
    return None


def _row_get(row: pd.Series, key: str, default: object = None) -> object:  # type: ignore[type-arg]
    """Accès sécurisé Series — évite l'ambiguïté str | NAType de Series.get() (Pylance-safe)."""
    try:
        val = row[key]
        if val is None:
            return default
        try:
            return default if pd.isna(val) else val  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return val
    except KeyError:
        return default


def _scalar(row: pd.Series, key: str, default: str = "") -> str:  # type: ignore[type-arg]
    try:
        val = row[key]
        return default if pd.isna(val) else str(val)  # type: ignore[arg-type]
    except (KeyError, TypeError):
        return default


def _get_float(row: pd.Series, key: str) -> float | None:  # type: ignore[type-arg]
    try:
        val = row[key]
        return None if pd.isna(val) else float(val)  # type: ignore[arg-type]
    except (KeyError, TypeError, ValueError):
        return None


def get_competitiveness(pct: float | None, params: dict[str, float]) -> str:
    if pct is None:
        return "N/A"
    if pct <= params["seuil_tres_competitif"]:
        return "✅ Très compétitif"
    if pct <= params["seuil_competitif"]:
        return "✅ Compétitif"
    if pct <= params["seuil_non_competitif"]:
        return "⚠️ Proche"
    return "❌ Non compétitif"


def build_reverse_maps(
    pension_config: dict[str, dict[str, list[str]]],
    cancel_config:  dict[str, dict[str, list[str]]],
) -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    orx_pension_rev = {v: n for n, vals in pension_config.items() for v in vals.get("orx", [])}
    bkg_meal_rev    = {v: n for n, vals in pension_config.items() for v in vals.get("bkg", [])}
    bkg_cancel_rev  = {v: n for n, vals in cancel_config.items()  for v in vals.get("bkg", [])}
    return orx_pension_rev, bkg_meal_rev, bkg_cancel_rev


# ── OPTIMISATION 2 : O(n) via set lookup ──────────────────────
def compute_date_coverage(df: pd.DataFrame) -> dict[str, object]:
    """Couverture des dates — O(n) grâce à un set de lookup."""
    key_cols    = ["Date de départ", "Nb nuits"]
    all_windows = df[key_cols].drop_duplicates().copy()

    covered_keys: set[tuple[object, object]] = set(
        zip(
            df.loc[df["Prix BKG (min)"].notna(), "Date de départ"],
            df.loc[df["Prix BKG (min)"].notna(), "Nb nuits"],
        )
    )
    all_windows["_has_bkg"] = [
        (d, n) in covered_keys
        for d, n in zip(all_windows["Date de départ"], all_windows["Nb nuits"])
    ]

    n_total   = len(all_windows)
    n_covered = int(all_windows["_has_bkg"].sum())
    rate      = n_covered / n_total if n_total else 0.0
    return {
        "total":   n_total,
        "covered": n_covered,
        "missing": n_total - n_covered,
        "rate":    rate,
        "detail":  all_windows.sort_values(key_cols),
    }


def build_bench_header(df: pd.DataFrame) -> tuple[str, str, str]:
    hotel_vals = df["Hotel"].replace("", pd.NA).dropna().unique()
    hotel_name = str(hotel_vals[0]) if len(hotel_vals) > 0 else "Hôtel inconnu"
    dates_parsed = pd.to_datetime(
        df["Date de départ"], format="%d/%m/%Y", errors="coerce"
    ).dropna()
    if not dates_parsed.empty:
        date_min   = dates_parsed.min().strftime("%d/%m/%Y")
        date_max   = dates_parsed.max().strftime("%d/%m/%Y")
        date_range = f"{date_min} → {date_max}" if date_min != date_max else date_min
    else:
        date_range = "—"
    nuits = sorted([int(n) for n in df["Nb nuits"].dropna().unique()])
    if len(nuits) == 0:
        nuits_str = "—"
    elif len(nuits) == 1:
        nuits_str = str(nuits[0])
    else:
        nuits_str = ", ".join(str(n) for n in nuits[:-1]) + " et " + str(nuits[-1])
    return hotel_name, date_range, nuits_str


def step_title(n: int, label: str) -> None:
    st.markdown(
        f'<div class="step-title"><span class="step-badge">{n}</span>{label}</div>',
        unsafe_allow_html=True,
    )
    st.markdown("")


def kpi(label: str, value: str, cls: str = "") -> None:
    st.markdown(
        f'<div class="kpi-row"><span class="kpi-label">{label}</span>'
        f'<span class="kpi-value {cls}">{value}</span></div>',
        unsafe_allow_html=True,
    )


def render_seuil_stats(df_all: pd.DataFrame, n_comparable: int) -> None:
    comp_counts = df_all["_competitivite"].value_counts()
    for label, _cls, bg in SEUIL_LABELS:
        count = int(comp_counts.get(label, 0))
        pct   = count / n_comparable if n_comparable else 0.0
        st.markdown(
            f'<div class="seuil-row" style="background:{bg};">'
            f'<span class="seuil-label">{label}</span>'
            f'<span><span class="seuil-count">{count:,}</span>'
            f'<span class="seuil-pct">({pct:.1%})</span></span>'
            f'</div>',
            unsafe_allow_html=True,
        )


def _fmt_price(v: object) -> str:
    if v is None or (not isinstance(v, str) and pd.isna(v)):  # type: ignore[arg-type]
        return "—"
    f = float(v)  # type: ignore[arg-type]
    return f"{f:,.0f}" if f == int(f) else f"{f:,.2f}"


def _fmt_ecart(v: object) -> str:
    if v is None or (not isinstance(v, str) and pd.isna(v)):  # type: ignore[arg-type]
        return "—"
    f = float(v)  # type: ignore[arg-type]
    return f"{f:+,.0f}" if f == int(f) else f"{f:+,.2f}"


def render_paginated_table(
    df_display: pd.DataFrame,
    df_full:    pd.DataFrame,
    page_key:      str,
    page_size_key: str,
) -> None:
    fc1, fc2, fc3 = st.columns([2, 2, 2])
    with fc1:
        hide_na = st.checkbox(
            "Masquer les lignes sans comparaison BKG (N/A)",
            value=False, key="filter_hide_na",
        )
    with fc2:
        if "Date de départ" in df_display.columns:
            raw_dates = df_display["Date de départ"].replace("", pd.NA).dropna().unique().tolist()
            sorted_dates = sorted(
                raw_dates,
                key=lambda d: pd.to_datetime(d, format="%d/%m/%Y", errors="coerce"),
            )
            date_options: list[str] = ["Toutes"] + sorted_dates
        else:
            date_options = ["Toutes"]
        selected_date = st.selectbox("Filtrer par date de départ", date_options, key="filter_date")
    with fc3:
        nuits_raw = sorted(
            df_display["Nb nuits"].dropna().unique().tolist()
        ) if "Nb nuits" in df_display.columns else []
        nuits_options: list[str] = ["Toutes"] + [str(int(n)) for n in nuits_raw]
        selected_nuits = st.selectbox("Filtrer par nb nuits", nuits_options, key="filter_nuits")

    mask = pd.Series([True] * len(df_display), index=df_display.index)
    if hide_na:
        mask &= df_full["_competitivite"] != "N/A"
    if selected_date != "Toutes" and "Date de départ" in df_display.columns:
        mask &= df_display["Date de départ"] == selected_date
    if selected_nuits != "Toutes" and "Nb nuits" in df_display.columns:
        try:
            nuits_val = int(selected_nuits)
            mask &= df_display["Nb nuits"].astype(float).astype(int) == nuits_val
        except (ValueError, TypeError):
            pass

    df_view      = df_display[mask]
    df_view_full = df_full[mask]
    n_rows       = len(df_view)

    ps_col, _, info_col = st.columns([2, 4, 2])
    with ps_col:
        page_size = st.selectbox("Lignes par page", PAGE_SIZES, index=1, key=page_size_key)
    with info_col:
        st.markdown(
            f"<div style='padding-top:28px;text-align:right;font-size:14px;color:#666;'>"
            f"{n_rows:,} ligne(s)</div>",
            unsafe_allow_html=True,
        )

    n_pages = max(1, (n_rows - 1) // page_size + 1)
    if page_key not in st.session_state:
        st.session_state[page_key] = 0
    if st.session_state[page_key] >= n_pages:
        st.session_state[page_key] = 0

    btn1, btn2, btn3 = st.columns([1, 4, 1])
    with btn1:
        if st.button("← Préc.", key=f"{page_key}_prev",
                     disabled=st.session_state[page_key] == 0):
            st.session_state[page_key] -= 1
            st.rerun()
    with btn2:
        st.markdown(
            f'<div class="pagination-bar">'
            f'Page <b>{st.session_state[page_key] + 1}</b> / <b>{n_pages}</b>'
            f'</div>',
            unsafe_allow_html=True,
        )
    with btn3:
        if st.button("Suiv. →", key=f"{page_key}_next",
                     disabled=st.session_state[page_key] == n_pages - 1):
            st.session_state[page_key] += 1
            st.rerun()

    start     = st.session_state[page_key] * page_size
    end       = start + page_size
    df_page   = df_view.iloc[start:end]
    df_p_full = df_view_full.iloc[start:end]

    def style_row(row: pd.Series) -> list[str]:  # type: ignore[type-arg]
        comp  = df_p_full.at[row.name, "_competitivite"]
        color = ROW_COLORS.get(str(comp), ("", ""))[1]
        return [f"background-color: {color}" if color else ""] * len(row)

    st.dataframe(
        df_page.style.apply(style_row, axis=1).format(
            {
                "Ecart PCT":          "{:.1%}",
                "Ecart EUR":          _fmt_ecart,
                "Prix BKG (min)":     _fmt_price,
                "Prix ORX / chambre": _fmt_price,
                "Prix de vente TTC":  _fmt_price,
            },
        ),
        use_container_width=True,
        hide_index=True,
        height=min(52 + page_size * 36, 720),
    )


@st.cache_data(show_spinner="Lecture du fichier...")
def load_data(file_bytes: bytes) -> tuple[pd.DataFrame, pd.DataFrame]:
    xls      = pd.ExcelFile(io.BytesIO(file_bytes))
    required = ["1. INPUT ORX EXPORT", "2. OUTPUT BKG SCRAP"]
    missing  = [s for s in required if s not in xls.sheet_names]
    if missing:
        raise ValueError(f"Feuilles manquantes : {', '.join(missing)}")
    df_o = pd.read_excel(xls, sheet_name="1. INPUT ORX EXPORT")
    df_b = pd.read_excel(xls, sheet_name="2. OUTPUT BKG SCRAP")
    df_o.columns = df_o.columns.str.strip()
    df_b.columns = df_b.columns.str.strip()
    return df_o, df_b


# ── OPTIMISATION 3 : build_excel au niveau module + cache ─────
@st.cache_data(show_spinner=False)
def build_excel(df_out: pd.DataFrame, orig_bytes: bytes, _df_full: pd.DataFrame) -> bytes:
    """
    Génère le fichier Excel avec mise en forme.
    Mis en cache : ne recalcule que si df_out ou orig_bytes changent.
    _df_full préfixé _ pour être ignoré par st.cache_data (non-hashable).
    """
    wb = openpyxl.load_workbook(io.BytesIO(orig_bytes))
    sn = "4. RAPPORT DETAILLE"
    if sn in wb.sheetnames:
        del wb[sn]
    ws       = wb.create_sheet(sn)
    headers  = list(df_out.columns)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin_bdr = Border(
        bottom=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin",  color="D0D0D0"),
    )
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font, cell.fill = hdr_font, hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(h, 18)
    ws.row_dimensions[1].height = 32

    for ri, (orig_idx, row_data) in enumerate(df_out.iterrows(), 2):
        comp     = str(_df_full.at[orig_idx, "_competitivite"])
        row_fill = PatternFill("solid", fgColor=ROW_COLORS.get(comp, ("F2F2F2", ""))[0])
        for ci, h in enumerate(headers, 1):
            val  = row_data[h]
            cell = ws.cell(row=ri, column=ci)
            cell.font   = Font(size=10)
            cell.border = thin_bdr
            cell.fill   = row_fill
            is_nan = val is None or (not isinstance(val, str) and pd.isna(val))
            if is_nan:
                cell.value = None
            elif h == "Ecart PCT" and isinstance(val, (int, float)):
                cell.value         = val
                cell.number_format = "0.0%"
                cell.alignment     = Alignment(horizontal="right")
            elif h in ("Ecart EUR", "Prix BKG (min)", "Prix ORX / chambre", "Prix de vente TTC") \
                    and isinstance(val, (int, float)):
                cell.value         = val
                cell.number_format = '#,##0.00 "EUR"'
                cell.alignment     = Alignment(horizontal="right")
            elif h in ("Date de départ", "Nb nuits"):
                cell.value     = val
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = val

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════
# ÉTAPE 0 — CHARGEMENT
# ══════════════════════════════════════════════════════════════

step_title(0, "Chargement du fichier Excel")

uploaded = st.file_uploader(
    "Fichier .xlsx — feuilles attendues : 1. INPUT ORX EXPORT  |  2. OUTPUT BKG SCRAP",
    type=["xlsx"],
)

if uploaded is None:
    st.info("👆 Chargez votre fichier Excel pour démarrer la configuration.")

else:
    file_bytes = uploaded.read()

    try:
        df_orx, df_bkg = load_data(file_bytes)
    except Exception as e:
        st.error(f"❌ Erreur de lecture : {e}")

    else:
        bkg_hotel_col: str = str(df_bkg.columns[0])

        orx_categories     = safe_unique(df_orx, "Catégorie")
        orx_pensions_raw   = safe_unique(df_orx, "Pension")
        bkg_room_types     = safe_unique(df_bkg, "Room Type")
        bkg_meal_plans_raw = safe_unique(df_bkg, "Meal Plan")
        bkg_cancel_raw     = safe_unique(df_bkg, "Cancellation Policy")

        c1, c2 = st.columns(2)
        c1.success(f"✅ ORX — {len(df_orx):,} lignes · {len(orx_categories)} catégorie(s) · {len(orx_pensions_raw)} libellé(s) pension")
        c2.success(f"✅ BKG — {len(df_bkg):,} lignes · {len(bkg_room_types)} type(s) de chambre · {len(bkg_cancel_raw)} politique(s) d'annulation")
        st.caption(f"🏨 Colonne hôtel (BKG col. 1) : **{bkg_hotel_col}**")
        st.markdown("---")

        tab_guide, tab_params, tab_rooms, tab_pensions, tab_cancel, tab_report = st.tabs([
            "📖  Guide",
            "⚙️  1 · Paramètres",
            "🛏️  2 · Chambres",
            "🍽️  3 · Pensions",
            "📋  4 · Annulation",
            "📊  5 · Rapport",
        ])

        # ── ONGLET GUIDE ──────────────────────────────────────
        with tab_guide:
            st.markdown("""
<div style="background:#EEF4FF;border-left:5px solid #1F3864;border-radius:8px;padding:20px 24px;margin-bottom:28px;">
    <div style="font-size:22px;font-weight:800;color:#1F3864;margin-bottom:6px;">🏨 Bienvenue sur le Benchmark Tarifaire</div>
    <div style="font-size:15px;color:#444;">
        Cet outil vous permet de comparer en quelques clics les prix de vente Orchestra
        avec les prix en temps réel scraped sur Booking.com — et d'identifier en un coup d'œil
        où vous êtes compétitif, et où vous ne l'êtes pas.
    </div>
</div>
""", unsafe_allow_html=True)

            st.markdown("## 🧭 Partie 1 — Comment ça marche ?")
            st.caption("Guide simple et rapide pour prendre l'outil en main.")
            st.markdown("")
            g1, g2 = st.columns(2, gap="large")

            with g1:
                st.markdown("""
<div style="background:#fff;border:1px solid #E0E0E0;border-radius:10px;padding:18px 20px;margin-bottom:16px;">
    <div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:8px;">📂 Étape 1 — Chargez votre fichier</div>
    <div style="font-size:14px;color:#444;line-height:1.7;">
        Glissez votre fichier <b>.xlsx</b> dans la zone de chargement.<br>
        Il doit contenir <b>deux feuilles</b> :
        <ul style="margin:6px 0 0 16px;">
            <li><b>1. INPUT ORX EXPORT</b> — votre export Orchestra</li>
            <li><b>2. OUTPUT BKG SCRAP</b> — le scraping Booking.com</li>
        </ul>
        L'outil détecte automatiquement les catégories, pensions et types de chambres disponibles.
    </div>
</div>
<div style="background:#fff;border:1px solid #E0E0E0;border-radius:10px;padding:18px 20px;margin-bottom:16px;">
    <div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:8px;">⚙️ Étape 2 — Configurez les seuils</div>
    <div style="font-size:14px;color:#444;line-height:1.7;">
        Dans l'onglet <b>Paramètres</b>, définissez vos seuils de compétitivité en %.<br>
        Les valeurs par défaut sont <b>-10 / -15 / -20 %</b> mais vous pouvez les ajuster selon votre stratégie tarifaire.
    </div>
</div>
<div style="background:#fff;border:1px solid #E0E0E0;border-radius:10px;padding:18px 20px;margin-bottom:16px;">
    <div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:8px;">🛏️ Étape 3 — Mappez les chambres</div>
    <div style="font-size:14px;color:#444;line-height:1.7;">
        Pour chaque <b>catégorie ORX</b>, sélectionnez les <b>types de chambre Booking</b> équivalents.<br>
        Sans ce mapping, aucune comparaison ne sera possible.
    </div>
</div>
""", unsafe_allow_html=True)

            with g2:
                st.markdown("""
<div style="background:#fff;border:1px solid #E0E0E0;border-radius:10px;padding:18px 20px;margin-bottom:16px;">
    <div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:8px;">🍽️ Étape 4 — Mappez les pensions</div>
    <div style="font-size:14px;color:#444;line-height:1.7;">
        Associez les libellés ORX et BKG à une valeur normalisée commune (ex. <i>PDJ, DP, PC, TI, LS</i>).<br>
        La comparaison n'est effectuée <b>qu'entre pensions de même type</b>.
    </div>
</div>
<div style="background:#fff;border:1px solid #E0E0E0;border-radius:10px;padding:18px 20px;margin-bottom:16px;">
    <div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:8px;">📋 Étape 5 — Mappez les annulations</div>
    <div style="font-size:14px;color:#444;line-height:1.7;">
        Normalisez les politiques d'annulation BKG (ex. <i>Free cancellation → AG</i>).<br>
        La politique ORX de référence sert uniquement à l'affichage dans le rapport.
    </div>
</div>
<div style="background:#fff;border:1px solid #E0E0E0;border-radius:10px;padding:18px 20px;margin-bottom:16px;">
    <div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:8px;">📊 Étape 6 — Générez le rapport</div>
    <div style="font-size:14px;color:#444;line-height:1.7;">
        Cliquez sur <b>Générer le rapport</b> dans l'onglet Rapport.<br>
        Vous obtenez instantanément :
        <ul style="margin:6px 0 0 16px;">
            <li>Un tableau coloré selon la compétitivité</li>
            <li>Des KPIs synthétiques (écarts moyens, taux de couverture…)</li>
            <li>Un fichier Excel exportable avec mise en forme automatique</li>
        </ul>
    </div>
</div>
""", unsafe_allow_html=True)

            st.markdown("""
<div style="background:#F0FFF4;border:1px solid #B7EBD0;border-radius:10px;padding:14px 20px;margin-top:4px;">
    <div style="font-size:14px;color:#375623;">
        💡 <b>Astuce</b> — Le rapport est <b>persistant</b> : naviguez entre les onglets sans perdre vos résultats.
        Recliquez sur <i>Générer</i> uniquement si vous modifiez un mapping ou un seuil.
    </div>
</div>
""", unsafe_allow_html=True)

            st.markdown("")
            st.markdown("#### 🎨 Lecture des couleurs")
            leg_data = [
                ("#C6EFCE", "✅ Très compétitif",  "Ecart ≤ −20 %",          "Prix nettement inférieur à Booking"),
                ("#E2EFDA", "✅ Compétitif",        "−20 % < Ecart ≤ −15 %",  "Bonne position tarifaire"),
                ("#FFEB9C", "⚠️ Proche",            "−15 % < Ecart ≤ −10 %",  "Vigilance — surveiller l'évolution"),
                ("#FFC7CE", "❌ Non compétitif",    "Ecart > −10 %",           "Prix ORX supérieur ou trop proche de Booking"),
                ("#F2F2F2", "⬜ N/A",               "—",                       "Aucune correspondance BKG trouvée"),
            ]
            st.dataframe(
                pd.DataFrame(leg_data, columns=["Couleur", "Indicateur", "Condition", "Interprétation"])
                .style.apply(
                    lambda row: [f"background-color:{row['Couleur']};font-weight:600"] * 4,  # type: ignore[return-value]
                    axis=1,
                ),
                hide_index=True,
                use_container_width=True,
            )

            st.markdown("---")
            st.markdown("## ⚙️ Partie 2 — Workflow des traitements")
            st.caption("Description technique du pipeline de normalisation et de comparaison.")
            st.markdown("")

            st.markdown("""
<div style="background:#F8F9FA;border:1px solid #DEE2E6;border-radius:10px;padding:20px 24px;margin-bottom:20px;">
<div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:12px;">① Chargement &amp; nettoyage des données</div>
<div style="font-size:14px;color:#444;line-height:1.8;">
    <b>Feuille ORX</b> — chaque ligne représente une offre Orchestra avec son prix TTC, sa catégorie,
    sa pension, son type de prix (1* = par personne, 2* = par chambre) et sa date de départ.<br>
    <b>Feuille BKG</b> — chaque ligne est un prix scraping Booking avec la date de check-in,
    le type de chambre, le meal plan, la politique d'annulation et le prix total pour 2 personnes.<br><br>
    Les en-têtes sont normalisés (<code>str.strip()</code>) et les dates parsées en objets <code>date</code> Python.
</div>
</div>
<div style="background:#F8F9FA;border:1px solid #DEE2E6;border-radius:10px;padding:20px 24px;margin-bottom:20px;">
<div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:12px;">② Normalisation des pensions</div>
<div style="font-size:14px;color:#444;line-height:1.8;">
    Chaque libellé brut ORX (ex. <i>"Petit déjeuner inclus"</i>) et BKG (ex. <i>"Breakfast included"</i>)
    est converti en valeur canonique : <code>PDJ</code>, <code>DP</code>, <code>PC</code>, <code>TI</code>, <code>LS</code>.<br><br>
    La pré-normalisation BKG est effectuée <b>une seule fois avant la boucle</b> (colonne <code>_meal_norm</code>).
    Les libellés non mappés sont flaggés <i>"Non mappé : …"</i>.
</div>
</div>
<div style="background:#F8F9FA;border:1px solid #DEE2E6;border-radius:10px;padding:20px 24px;margin-bottom:20px;">
<div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:12px;">③ Normalisation du prix ORX → prix chambre</div>
<div style="font-size:14px;color:#444;line-height:1.8;">
    Les prix BKG sont toujours <b>par chambre pour 2 personnes</b>.<br>
    <ul style="margin:8px 0 0 18px;">
        <li><b>1*</b> (prix par personne) → × 2 pour obtenir le prix chambre</li>
        <li><b>2*</b> (prix par chambre) → utilisé tel quel</li>
    </ul>
</div>
</div>
<div style="background:#F8F9FA;border:1px solid #DEE2E6;border-radius:10px;padding:20px 24px;margin-bottom:20px;">
<div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:12px;">④ Matching ORX ↔ BKG</div>
<div style="font-size:14px;color:#444;line-height:1.8;">
    Filtre séquentiel strict via un <b>index dict O(1)</b> pré-calculé avant la boucle :
    <ol style="margin:8px 0 0 18px;line-height:2.0;">
        <li><b>Date exacte</b> — <code>Date de départ ORX == Check-in BKG</code></li>
        <li><b>Type de chambre mappé</b> — Room Type BKG dans la liste associée à la catégorie ORX</li>
        <li><b>Même pension normalisée</b> — <code>_meal_norm BKG == pension_norm ORX</code></li>
    </ol>
    Lorsque plusieurs BKG correspondent, le <b>prix minimum</b> est retenu.
</div>
</div>
<div style="background:#F8F9FA;border:1px solid #DEE2E6;border-radius:10px;padding:20px 24px;margin-bottom:20px;">
<div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:12px;">⑤ Calcul des écarts &amp; scoring</div>
<div style="font-size:14px;color:#444;line-height:1.8;">
    <div style="background:#fff;border:1px solid #e0e0e0;border-radius:6px;padding:12px 16px;font-family:monospace;font-size:14px;">
        Ecart EUR = Prix ORX/chambre − Prix BKG<br>
        Ecart %   = (Prix ORX/chambre − Prix BKG) / Prix BKG
    </div>
    <br>L'Ecart % est comparé aux seuils pour attribuer l'un des 4 niveaux de compétitivité.
</div>
</div>
<div style="background:#F8F9FA;border:1px solid #DEE2E6;border-radius:10px;padding:20px 24px;">
<div style="font-size:16px;font-weight:700;color:#1F3864;margin-bottom:12px;">⑥ Production du rapport</div>
<div style="font-size:14px;color:#444;line-height:1.8;">
    <ul style="margin:8px 0 0 18px;line-height:2.0;">
        <li><b>KPIs synthétiques</b> — écarts moyens/min/max, répartition par niveau, taux de couverture</li>
        <li><b>Tableau interactif</b> — filtrable par date, par nb nuits, masquage N/A, paginé</li>
        <li><b>Export Excel</b> — feuille <code>4. RAPPORT DETAILLE</code> avec coloration conditionnelle et filtres auto</li>
    </ul>
</div>
</div>
""", unsafe_allow_html=True)

        # ── ONGLET 1 ──────────────────────────────────────────
        with tab_params:
            step_title(1, "Paramètres de benchmark")
            col_a, col_b = st.columns([1, 1], gap="large")
            with col_a:
                st.markdown("**Politique d'annulation de référence ORX**")
                ref_policy = st.selectbox(
                    "ref_policy", options=ANNULATIONS_NORM,
                    index=ANNULATIONS_NORM.index("RF - Remboursable"),
                    label_visibility="collapsed",
                )
                st.info(f"Référence ORX : **{ref_policy}**")
            with col_b:
                st.markdown("**Seuils de compétitivité** *(valeurs négatives en %)*")
                st.caption("Ecart = (Prix ORX - Prix BKG) / Prix BKG")
                seuil_proche = st.number_input(
                    "Seuil Non compétitif / Proche",
                    value=-10, min_value=-99, max_value=0, step=1, format="%d",
                ) / 100
                seuil_competitif = st.number_input(
                    "Seuil Proche / Compétitif",
                    value=-15, min_value=-99, max_value=0, step=1, format="%d",
                ) / 100
                seuil_tres_competitif = st.number_input(
                    "Seuil Compétitif / Très compétitif",
                    value=-20, min_value=-99, max_value=0, step=1, format="%d",
                ) / 100
            st.markdown("---")
            st.dataframe(pd.DataFrame([
                {"Indicateur": "❌ Non compétitif",  "Condition": f"Ecart > {int(seuil_proche*100)}%"},
                {"Indicateur": "⚠️ Proche",          "Condition": f"{int(seuil_competitif*100)}% < Ecart <= {int(seuil_proche*100)}%"},
                {"Indicateur": "✅ Compétitif",       "Condition": f"{int(seuil_tres_competitif*100)}% < Ecart <= {int(seuil_competitif*100)}%"},
                {"Indicateur": "✅ Très compétitif",  "Condition": f"Ecart <= {int(seuil_tres_competitif*100)}%"},
            ]), hide_index=True, use_container_width=True)
            params: dict[str, float] = {
                "seuil_non_competitif":  float(seuil_proche),
                "seuil_competitif":      float(seuil_competitif),
                "seuil_tres_competitif": float(seuil_tres_competitif),
            }

        # ── ONGLET 2 ──────────────────────────────────────────
        with tab_rooms:
            step_title(2, "Mapping — Types de chambres")
            st.caption("Pour chaque catégorie ORX, associez un ou plusieurs types de chambres Booking.")
            st.markdown("")
            room_mapping: dict[str, list[str]] = {}
            for cat in orx_categories:
                room_mapping[cat] = st.multiselect(
                    label=f"**{cat}**",
                    options=bkg_room_types,
                    key=f"room__{cat}",
                    placeholder="Types de chambre BKG correspondants...",
                )
            preview = [{"Catégorie ORX": c, "Room Type BKG": rt}
                       for c, rts in room_mapping.items() for rt in rts]
            if preview:
                st.markdown("---")
                st.dataframe(pd.DataFrame(preview), hide_index=True, use_container_width=True)
            else:
                st.warning("⚠️ Aucun mapping chambre défini — toutes les lignes seront N/A.")

        # ── ONGLET 3 ──────────────────────────────────────────
        with tab_pensions:
            step_title(3, "Mapping — Pensions / Meal Plans")
            st.caption(
                "Pour chaque valeur normalisée, associez les libellés ORX ET les libellés BKG. "
                "La comparaison s'effectuera uniquement entre pensions de même type normalisé."
            )
            st.markdown("")
            pension_config: dict[str, dict[str, list[str]]] = {}
            for norm in PENSIONS_NORM:
                with st.expander(f"🍽️  **{norm}**", expanded=False):
                    col_orx, col_bkg = st.columns(2, gap="medium")
                    with col_orx:
                        st.markdown("**Sources ORX** *(colonne : Pension)*")
                        orx_vals = st.multiselect(
                            "orx", options=orx_pensions_raw, key=f"p_orx__{norm}",
                            label_visibility="collapsed", placeholder="Libellés ORX...",
                        )
                    with col_bkg:
                        st.markdown("**Sources BKG** *(colonne : Meal Plan)*")
                        bkg_vals = st.multiselect(
                            "bkg", options=bkg_meal_plans_raw, key=f"p_bkg__{norm}",
                            label_visibility="collapsed", placeholder="Libellés BKG...",
                        )
                    pension_config[norm] = {"orx": orx_vals, "bkg": bkg_vals}
            orx_p_rev, bkg_m_rev, _ = build_reverse_maps(pension_config, {})
            unmapped_orx_p = [v for v in orx_pensions_raw   if v not in orx_p_rev]
            unmapped_bkg_m = [v for v in bkg_meal_plans_raw if v not in bkg_m_rev]
            if unmapped_orx_p:
                st.warning(f"⚠️ Libellés ORX non mappés : {', '.join(unmapped_orx_p)}")
            if unmapped_bkg_m:
                st.warning(f"⚠️ Libellés BKG non mappés : {', '.join(unmapped_bkg_m)}")

        # ── ONGLET 4 ──────────────────────────────────────────
        with tab_cancel:
            step_title(4, "Mapping — Politiques d'annulation")
            st.markdown("")
            cancel_config: dict[str, dict[str, list[str]]] = {}
            for norm in ANNULATIONS_NORM:
                with st.expander(f"📋  **{norm}**", expanded=False):
                    st.markdown("**Sources BKG** *(colonne : Cancellation Policy)*")
                    bkg_cancel_vals = st.multiselect(
                        "bkg", options=bkg_cancel_raw, key=f"c_bkg__{norm}",
                        label_visibility="collapsed", placeholder="Libellés BKG...",
                    )
                    cancel_config[norm] = {"bkg": bkg_cancel_vals}
            _, _, bkg_c_rev = build_reverse_maps(pension_config, cancel_config)
            unmapped_bkg_c = [v for v in bkg_cancel_raw if v not in bkg_c_rev]
            if unmapped_bkg_c:
                st.warning(f"⚠️ Politiques BKG non mappées : {', '.join(unmapped_bkg_c)}")

        # ── ONGLET 5 — RAPPORT ────────────────────────────────
        with tab_report:
            step_title(5, "Rapport de compétitivité")
            missing_room_cats = [c for c, v in room_mapping.items() if not v]
            if missing_room_cats:
                st.warning(f"⚠️ Catégories sans mapping chambre : **{', '.join(missing_room_cats)}**")
            st.markdown("")
            generate_btn = st.button("🚀 Générer le rapport", type="primary", use_container_width=True)

            if generate_btn:
                orx_pension_rev, bkg_meal_rev, bkg_cancel_rev = build_reverse_maps(
                    pension_config, cancel_config
                )
                with st.spinner("Traitement en cours..."):
                    df_orx_p = df_orx.copy()
                    df_bkg_p = df_bkg.copy()
                    df_orx_p["Date de départ"] = pd.to_datetime(
                        df_orx_p["Date de départ"], errors="coerce"
                    ).dt.date
                    df_bkg_p["Check-in"] = pd.to_datetime(
                        df_bkg_p["Check-in"], errors="coerce"
                    ).dt.date

                    # Pré-normalisation meal plans BKG
                    df_bkg_p["_meal_norm"] = (
                        df_bkg_p["Meal Plan"].astype(str).str.strip()
                        .map(lambda m: bkg_meal_rev.get(str(m), ""))
                    )

                    # ── OPTIMISATION 1 : index BKG O(1) ──────
                    # Pour chaque (date, meal_norm, room_type) → ligne au prix minimum
                    bkg_min_df = (
                        df_bkg_p.sort_values("Price")
                        .groupby(["Check-in", "_meal_norm", "Room Type"], as_index=False)
                        .first()
                    )
                    bkg_lookup: dict[tuple[object, str, str], pd.Series] = {  # type: ignore[type-arg]
                        (row["Check-in"], str(row["_meal_norm"]), str(row["Room Type"])): row
                        for _, row in bkg_min_df.iterrows()
                    }

                    results: list[dict[str, object]] = []
                    n_no_map = 0
                    n_no_bkg = 0

                    for _, orx_row in df_orx_p.iterrows():
                        date_dep    = _row_get(orx_row, "Date de départ")
                        nb_nuits    = _row_get(orx_row, "Nb nuits")
                        categorie   = str(_row_get(orx_row, "Catégorie",   "") or "").strip()
                        pension_raw = str(_row_get(orx_row, "Pension",     "") or "").strip()

                        pension_norm       = orx_pension_rev.get(pension_raw, f"Non mappé : {pension_raw}")
                        pension_norm_valid = bool(pension_norm and not pension_norm.startswith("Non mappé"))

                        try:
                            date_label = date_dep.strftime("%d/%m/%Y") if date_dep else ""  # type: ignore[union-attr]
                        except Exception:
                            date_label = str(date_dep) if date_dep else ""

                        prix_ttc_raw     = _row_get(orx_row, "Prix de vente TTC")
                        prix_orx_num     = normalize_price(prix_ttc_raw)
                        prix_orx_chambre = round(prix_orx_num * 2, 2) if prix_orx_num is not None else None
                        room_types       = room_mapping.get(categorie, [])

                        if not room_types:
                            n_no_map += 1
                            nom_hotel = ""
                            bkg_room = bkg_meal = bkg_meal_norm = bkg_cancel_norm = "N/A"
                            prix_bkg: float | None  = None
                            ecart_eur: float | None = None
                            ecart_pct: float | None = None
                            competitivite = "N/A"
                        else:
                            # ── Lookup O(1) via dict pré-indexé ──
                            meal_key      = pension_norm if pension_norm_valid else ""
                            cheapest_row_found: pd.Series | None = None  # type: ignore[type-arg]
                            cheapest_price_found = float("inf")

                            for rt in room_types:
                                candidate = bkg_lookup.get((date_dep, meal_key, str(rt)))
                                if candidate is not None:
                                    p = _get_float(candidate, "Price")
                                    if p is not None and p < cheapest_price_found:
                                        cheapest_price_found = p
                                        cheapest_row_found = candidate

                            if cheapest_row_found is None:
                                n_no_bkg += 1
                                nom_hotel = ""
                                bkg_room = bkg_meal = bkg_meal_norm = bkg_cancel_norm = "N/A"
                                prix_bkg = ecart_eur = ecart_pct = None
                                competitivite = "N/A"
                            else:
                                cheapest_row       = cheapest_row_found
                                nom_hotel          = _scalar(cheapest_row, bkg_hotel_col).strip()
                                bkg_room           = _scalar(cheapest_row, "Room Type")
                                bkg_meal           = _scalar(cheapest_row, "Meal Plan")
                                bkg_meal_norm      = bkg_meal_rev.get(bkg_meal.strip(), f"Non mappé : {bkg_meal}")
                                bkg_cancel_raw_val = _scalar(cheapest_row, "Cancellation Policy")
                                bkg_cancel_norm    = bkg_cancel_rev.get(
                                    bkg_cancel_raw_val, f"Non mappé : {bkg_cancel_raw_val}"
                                )
                                prix_bkg = _get_float(cheapest_row, "Price")
                                if prix_orx_chambre is not None and prix_bkg is not None:
                                    ecart_eur     = round(prix_orx_chambre - prix_bkg, 2)
                                    ecart_pct     = round((prix_orx_chambre - prix_bkg) / prix_bkg, 4)
                                    competitivite = get_competitiveness(ecart_pct, params)
                                else:
                                    ecart_eur = ecart_pct = None
                                    competitivite = "N/A"

                        results.append({
                            "Hotel":                               nom_hotel,
                            "Date de départ":                      date_label,
                            "Nb nuits":                            nb_nuits,
                            "Catégorie ORX":                       categorie,
                            "Pension ORX (norm.)":                 pension_norm,
                            "Politique annulation ORX (réf.)":     ref_policy,
                            "Type de prix":                        _row_get(orx_row, "Type de prix"),
                            "Prix de vente TTC":                   prix_ttc_raw,
                            "Prix ORX / chambre":                  prix_orx_chambre,
                            "Room Type BKG":                       bkg_room,
                            "Meal Plan BKG":                       bkg_meal,
                            "Meal Plan BKG (norm.)":               bkg_meal_norm,
                            "Politique annulation BKG (norm.)":    bkg_cancel_norm,
                            "Prix BKG (min)":                      prix_bkg,
                            "Ecart EUR":                           ecart_eur,
                            "Ecart PCT":                           ecart_pct,
                            "_competitivite":                      competitivite,
                        })

                    st.session_state["df_rapport"] = pd.DataFrame(results)
                    st.session_state["n_no_map"]   = n_no_map
                    st.session_state["n_no_bkg"]   = n_no_bkg
                    st.session_state["file_bytes"] = file_bytes

            if "df_rapport" not in st.session_state:
                st.info("Cliquez sur **Générer le rapport** pour lancer l'analyse.")
            else:
                df   : pd.DataFrame = st.session_state["df_rapport"]
                n_nm : int          = st.session_state["n_no_map"]
                n_nb : int          = st.session_state["n_no_bkg"]

                df_comp      = df[df["Prix BKG (min)"].notna() & df["Ecart EUR"].notna()]
                df_scored    = df[df["_competitivite"] != "N/A"]
                n_total      = len(df)
                n_reussie    = len(df_comp)
                n_comparable = len(df_scored)
                taux         = n_reussie / n_total if n_total else 0.0

                hotel_name, date_range, nuits_str = build_bench_header(df)
                st.markdown("---")
                st.markdown(
                    f'<div class="bench-title">Synthèse de bench — {hotel_name}</div>'
                    f'<div class="bench-subtitle">'
                    f'Date de départ : {date_range} &nbsp;|&nbsp; Nuitées : {nuits_str}'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                sc1, sc2, sc3 = st.columns(3, gap="medium")
                with sc1:
                    st.markdown('<div class="kpi-section-title">Positionnement Orchestra</div>', unsafe_allow_html=True)
                    if not df_scored.empty:
                        render_seuil_stats(df_scored, n_comparable)
                        st.markdown(
                            f"<div style='font-size:13px;color:#666;margin-top:7px;'>"
                            f"Sur {n_comparable:,} comparaison(s) avec seuil calculé</div>",
                            unsafe_allow_html=True,
                        )
                    else:
                        st.caption("Aucune comparaison disponible.")

                with sc2:
                    st.markdown('<div class="kpi-section-title">Analyse des écarts</div>', unsafe_allow_html=True)
                    if not df_comp.empty:
                        def cls_e(v: float) -> str:
                            return "success" if v < 0 else "danger" if v > 0 else ""
                        kpi("Ecart moyen (EUR)",   f"{df_comp['Ecart EUR'].mean():+,.0f} EUR", cls_e(float(df_comp['Ecart EUR'].mean())))
                        kpi("Ecart moyen (%)",     f"{df_comp['Ecart PCT'].mean():+.1%}",      cls_e(float(df_comp['Ecart PCT'].mean())))
                        kpi("Ecart maximum (EUR)", f"{df_comp['Ecart EUR'].max():+,.0f} EUR",  cls_e(float(df_comp['Ecart EUR'].max())))
                        kpi("Ecart minimum (EUR)", f"{df_comp['Ecart EUR'].min():+,.0f} EUR",  cls_e(float(df_comp['Ecart EUR'].min())))
                    else:
                        st.caption("Aucune comparaison disponible.")

                with sc3:
                    st.markdown('<div class="kpi-section-title">Indicateurs généraux</div>', unsafe_allow_html=True)
                    kpi("Nombre total de lignes",        f"{n_total:,}")
                    kpi("Comparaisons réussies",          f"{n_reussie:,}")
                    kpi("Taux de comparaisons réussies",  f"{taux:.0%}",
                        "success" if taux >= 0.5 else "warning" if taux > 0 else "danger")
                    kpi("Lignes sans mapping chambre",    f"{n_nm:,}", "warning" if n_nm > 0 else "")
                    kpi("Lignes sans prix BKG",           f"{n_nb:,}", "warning" if n_nb > 0 else "")

                st.markdown("")
                tw        = compute_date_coverage(df)
                rate      = float(tw["rate"])  # type: ignore[arg-type]
                bar_color = "#375623" if rate >= 0.8 else "#9C6500" if rate >= 0.4 else "#C00000"
                bar_w     = int(rate * 100)

                with st.container(border=True):
                    st.markdown('<div class="kpi-section-title">Couverture des dates de départ</div>', unsafe_allow_html=True)
                    st.caption("Pour chaque combinaison (Date de départ x Nb nuits), vérifie si une correspondance BKG de même pension existe.")
                    tw1, tw2, tw3 = st.columns(3)
                    tw1.metric("Combinaisons totales",    str(tw["total"]))
                    tw2.metric("Avec correspondance BKG", str(tw["covered"]))
                    missing_val = int(tw["missing"])  # type: ignore[arg-type]
                    tw3.metric(
                        "Sans correspondance", str(missing_val),
                        delta=f"-{missing_val}" if missing_val > 0 else None,
                        delta_color="inverse",
                    )
                    st.markdown(
                        f'<div style="margin:8px 0 4px 0;font-size:14px;font-weight:600;color:{bar_color};">'
                        f'Taux de couverture : {rate:.1%}</div>'
                        f'<div class="tw-bar-wrap"><div class="tw-bar-fill" '
                        f'style="width:{bar_w}%;background:{bar_color};"></div></div>',
                        unsafe_allow_html=True,
                    )
                    with st.expander("Détail par date de départ", expanded=False):
                        detail = tw["detail"].copy()  # type: ignore[union-attr]
                        detail["Statut"] = detail["_has_bkg"].map(  # type: ignore[index]
                            {True: "✅ Couverte", False: "❌ Manquante"}
                        )
                        st.dataframe(
                            detail[["Date de départ", "Nb nuits", "Statut"]],  # type: ignore[index]
                            hide_index=True, use_container_width=True,
                        )

                st.markdown("---")
                legend_items = [
                    ("#C6EFCE", "Très compétitif"), ("#E2EFDA", "Compétitif"),
                    ("#FFEB9C", "Proche"),           ("#FFC7CE", "Non compétitif"),
                    ("#F2F2F2", "N/A"),
                ]
                st.markdown(
                    '<div class="legend-wrap">' +
                    "".join(
                        f'<div class="legend-item">'
                        f'<div class="legend-dot" style="background:{c};"></div>{lbl}'
                        f'</div>'
                        for c, lbl in legend_items
                    ) + "</div>",
                    unsafe_allow_html=True,
                )

                df_display = df.drop(columns=["_competitivite", "Hotel"])
                render_paginated_table(df_display, df, "report_page", "report_page_size")

                st.markdown("")
                excel_bytes = build_excel(df_display, st.session_state["file_bytes"], df)
                st.download_button(
                    label="⬇️  Télécharger le rapport Excel",
                    data=excel_bytes,
                    file_name="benchmark_rapport.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

# ══════════════════════════════════════════════════════════════
# FOOTER
# ══════════════════════════════════════════════════════════════

st.markdown(
    '<div class="app-footer">'
    '🏨 Benchmark Tarifaire ORX vs Booking.com &nbsp;·&nbsp; '
    '<strong>Salah CHERKAOUI</strong> &nbsp;·&nbsp; Equipe projet Voyage &amp; Loisirs'
    '</div>',
    unsafe_allow_html=True,
)